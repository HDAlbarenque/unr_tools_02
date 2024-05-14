# Graba, en BD de Firestore (cuenta uniregm@gmail.com), los metadatos de las munis
# extraidos de planilla Excel

import openpyxl

import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore


class metadatos:

    # Importa un Excel, y lo transforma en diccionario
    def excel2lista(self, archivo_excel: str):
        # Abrir el documento
        workbook = openpyxl.load_workbook(archivo_excel)

        # Seleccionar la hoja activa
        hoja = workbook.active

        # Obtener la última fila y columna con datos de la hoja activa
        ult_fila = hoja.max_row
        ult_col = hoja.max_column

        # Obtener la primera fila, como claves para el diccionario
        header = [
            "muni_nro",
            "muni_nombre",
            "Dia_descarga",
            "Mega_Carpeta",
            "Link_Acceso",
            "Usuario",
            "Password",
            "Ruta_Local",
        ]

        # Inicializar lista contenedora
        muni_filas = []

        # Iterar sobre las filas con datos, y agregarlas al diccionario
        for fila in hoja.iter_rows(
            min_row=2,
            max_row=ult_fila,
            min_col=1,
            max_col=ult_col,
            values_only=True,
        ):
            fila_datos = dict(zip(header, fila))
            muni_filas.append(fila_datos)

        return muni_filas

    # Función para grabar la lista obtenida con cosmos2lista en Firestore
    def lista2Firestore(self, muni_filas):
        if not firebase_admin._apps:
            cred = credentials.Certificate("DB/Firebase/fb_unire.json")
            firebase_admin.initialize_app(cred)

        db = firestore.client()

        # Grabar datos en la colección 'UNIRE'
        db.collection("Datos_Munis").document("Mega_metadatos").set(
            {"Datos_Backup": muni_filas}
        )


if __name__ == "__main__":
    import datetime

    print(f"Inicio proceso: {datetime.datetime.now().strftime('%H:%M:%S')}")

    arch_excel = "DB\Firebase\metadatos2fb.xlsx"
    oDatos = metadatos()
    metadatos = oDatos.excel2lista(archivo_excel=arch_excel)
    oDatos.lista2Firestore(muni_filas=metadatos)
    print(f"Hora de finalización: {datetime.datetime.now().strftime('%H:%M:%S')}")
    pass
